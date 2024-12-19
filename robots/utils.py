from collections import defaultdict

import openpyxl
from django.db.models import Count

from constants import HEADERS_EXCEL, ERROR_ROBOTS_LAST_WEEK
from .models import Robot


def fetch_robot_data(last_week):
    """Получение данных из базы."""
    robots = (
        Robot.objects
        .filter(created__gte=last_week)
        .values('model', 'version')
        .annotate(count=Count('id'))
    )
    models_data = defaultdict(list)

    for robot in robots:
        model = robot['model']
        version = robot['version']
        count = robot['count']
        models_data[model].append((version, count))

    return models_data


def create_excel_file(models_data):
    """Создание Excel-файла с данными."""
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    if not models_data:
        ws = wb.create_sheet(title='No Data')
        ws.append([ERROR_ROBOTS_LAST_WEEK])

    else:
        for model, versions in models_data.items():
            ws = wb.create_sheet(title=model)
            ws.append(HEADERS_EXCEL)
            for version, count in versions:
                ws.append([model, version, count])

    return wb
